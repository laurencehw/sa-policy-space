"""
SA Policy Space — Consolidated Expenditure Updater
====================================================
Fetches consolidated government expenditure by function from National Treasury
Budget Review pivot tables and upserts into the Supabase consolidated_expenditure table.

Source: National Treasury Budget Review — Consolidated account Pivot (Excel)
URL pattern: https://www.treasury.gov.za/documents/National%20Budget/{YEAR}/review/Budget%20{YEAR}%20-%20Consolidated%20account%20Pivot.xlsx

Values are in R thousands as published.

Usage:
    python update_consolidated.py                    # fetch latest (2026)
    python update_consolidated.py --year 2025        # specific budget year
    python update_consolidated.py --all-years        # fetch all available years
    python update_consolidated.py --dry-run          # show what would be fetched

Environment variables required:
    SUPABASE_URL      - Your Supabase project URL
    SUPABASE_KEY      - Your Supabase service_role key
"""

import os
import sys
import io
import logging
import argparse
import requests
import openpyxl
from supabase import create_client

AVAILABLE_YEARS = [2020, 2021, 2022, 2023, 2024, 2025, 2026]

TREASURY_URL_TEMPLATE = (
    "https://www.treasury.gov.za/documents/National%20Budget/{year}/review/"
    "Budget%20{year}%20-%20Consolidated%20account%20Pivot.xlsx"
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)


def download_pivot(year: int) -> list[dict]:
    """Download and parse the consolidated account pivot Excel file."""
    url = TREASURY_URL_TEMPLATE.format(year=year)
    log.info(f"  Fetching: {url}")
    resp = requests.get(url, timeout=120)
    resp.raise_for_status()
    log.info(f"  Downloaded {len(resp.content)} bytes")

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
    if "Data" not in wb.sheetnames:
        log.error(f"  No 'Data' sheet found in pivot file (sheets: {wb.sheetnames})")
        return []

    ws = wb["Data"]
    rows = []
    header = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c).strip() if c else "" for c in row]
            log.info(f"  Columns: {header}")
            continue

        record = dict(zip(header, row))
        budget_year = str(record.get("BudgetYear", "")).strip()
        function_group = str(record.get("Function group", "")).strip()
        econ2 = str(record.get("Econ2", "")).strip()
        econ3 = str(record.get("Econ3", "")).strip()
        financial_year = str(record.get("FinYear", "")).strip()
        value = record.get("Value")

        if not function_group or not financial_year:
            continue

        try:
            amount = float(value) if value is not None else None
        except (ValueError, TypeError):
            amount = None

        rows.append({
            "budget_year": budget_year,
            "function_group": function_group,
            "econ2": econ2,
            "econ3": econ3,
            "financial_year": financial_year,
            "amount_rthousands": amount,
        })

    wb.close()
    log.info(f"  Parsed {len(rows)} data rows")
    return rows


def get_supabase_client():
    url = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY") or os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY")
    if not url or not key:
        log.error("Missing SUPABASE_URL or SUPABASE_KEY environment variables")
        sys.exit(1)
    return create_client(url, key)


def upsert_rows(supabase, rows: list[dict], budget_year: str) -> int:
    """Delete existing rows for this budget year, then insert fresh data."""
    if not budget_year:
        log.error("  No budget_year — refusing to delete")
        return 0

    try:
        supabase.table("consolidated_expenditure").delete().eq(
            "budget_year", budget_year
        ).execute()
        log.info(f"  Deleted existing rows for {budget_year}")
    except Exception as e:
        log.error(f"  Error deleting rows: {e}")
        return 0

    BATCH_SIZE = 500
    total = 0
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i:i + BATCH_SIZE]
        try:
            supabase.table("consolidated_expenditure").insert(batch).execute()
            total += len(batch)
            log.info(f"  Inserted batch {i // BATCH_SIZE + 1}: {len(batch)} rows (total: {total})")
        except Exception as e:
            log.error(f"  Error inserting batch: {e}")
    return total


def main():
    parser = argparse.ArgumentParser(description="Update consolidated expenditure from Treasury pivot tables")
    parser.add_argument("--year", type=int, default=None, help="Budget year (e.g. 2026)")
    parser.add_argument("--all-years", action="store_true", help="Fetch all available years")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be fetched")
    args = parser.parse_args()

    if args.all_years:
        years = AVAILABLE_YEARS
    elif args.year:
        years = [args.year]
    else:
        years = [AVAILABLE_YEARS[-1]]

    log.info(f"Will fetch consolidated expenditure for budget years: {years}")

    if not args.dry_run:
        supabase = get_supabase_client()

    total_rows = 0

    for year in years:
        log.info(f"\n{'=' * 60}")
        log.info(f"Processing Budget {year}")

        try:
            rows = download_pivot(year)
        except Exception as e:
            log.error(f"  Failed to download/parse pivot: {e}")
            continue

        if not rows:
            log.warning(f"  No data rows extracted, skipping")
            continue

        if args.dry_run:
            funcs = set(r["function_group"] for r in rows)
            fys = sorted(set(r["financial_year"] for r in rows))
            log.info(f"  [DRY RUN] {len(rows)} rows")
            log.info(f"  Functions: {sorted(funcs)}")
            log.info(f"  Financial years: {fys}")
            for func in sorted(funcs):
                func_rows = [r for r in rows if r["function_group"] == func]
                total = sum(r["amount_rthousands"] or 0 for r in func_rows)
                log.info(f"    {func}: {len(func_rows)} rows, total R{total/1_000_000:.1f}bn")
            total_rows += len(rows)
            continue

        budget_year = rows[0]["budget_year"]
        inserted = upsert_rows(supabase, rows, budget_year)
        total_rows += inserted

    log.info(f"\n{'=' * 60}")
    log.info(f"Done! Processed {total_rows} rows across {len(years)} budget year(s).")
    if args.dry_run:
        log.info("(Dry run — nothing written to database)")


if __name__ == "__main__":
    main()
